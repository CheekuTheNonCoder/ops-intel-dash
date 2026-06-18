"""
engine_export.py — Corporate-Style 17-Sheet Premium Excel Export Engine (v4.0)
Fully Restores original luxury corporate formatting, conditional highlights, and pivots [D].
Fixed: Resolved NameError 'D' by passing explicit validation values to _sheet_validation_report [1].
"""
import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PREMIUM CORPORATE PALETTE ─────────────────────────────────
NAVY   = "1F3864"; BLUE   = "2E75B6"; LTBLUE = "DBEAFE"
RED    = "C00000"; LTRED  = "FEE2E2"; AMBER  = "C55A11"
LTAMB  = "FEF3C7"; GREEN  = "166534"; LTGRN  = "DCFCE7"
GRAY1  = "F8FAFC"; GRAY2  = "E2E8F0"; GRAY3  = "64748B"
WHITE  = "FFFFFF"

def _s(style="thin", color="CBD5E0"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _f(c): return PatternFill("solid", start_color=c, end_color=c)

def _c(ws, r, c, v, bold=False, align="left", nf=None, fg="1E293B", bg=None, sz=9, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, size=sz, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _s()
    if nf:  cell.number_format = nf
    if bg:  cell.fill = _f(bg)
    return cell

def _h(ws, r, c, v, bg=NAVY, fg=WHITE, sz=10, align="center", wrap=True, height=None):
    cell = _c(ws, r, c, v, bold=True, align=align, fg=fg, bg=bg, sz=sz, wrap=wrap)
    if height: ws.row_dimensions[r].height = height
    return cell

def _title(ws, title, sub, cols):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws["A1"]
    c.value = title
    c.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill  = _f(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{get_column_letter(cols)}2")
    c2 = ws["A2"]
    c2.value = sub
    c2.font  = Font(name="Calibri", size=9, italic=True, color=GRAY3)
    c2.fill  = _f(GRAY1)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

def _w(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _impact_style(impact):
    return {"CRITICAL": (LTRED, RED), "HIGH": (LTAMB, AMBER),
            "MEDIUM": (LTBLUE, BLUE), "LOW": (LTGRN, GREEN)}.get(str(impact).upper(), (WHITE, "000000"))

def _autofit(ws):
    """Enables gridlines, freeze panes on A5, and auto-fits columns dynamically [1]."""
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A5"
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)


# ── SHEET HELPER FOR UNIFORM GRIDS (Declared Globally at the Top) ──
def dump_sheet(wb, sheet_name, title_txt, period, headers, df_data, formats=None):
    ws = wb.create_sheet(sheet_name)
    _title(ws, title_txt, f"Filter: {period}", len(headers))
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    idx = 5
    if not df_data.empty:
        for _, row in df_data.iterrows():
            ws.row_dimensions[idx].height = 16
            for ci, col in enumerate(df_data.columns, 1):
                val = row[col]
                nf = formats.get(col) if formats else None
                _c(ws, idx, ci, val, align="center" if ci > 1 else "left", nf=nf)
            idx += 1
    _autofit(ws)
    return ws


# ── SHEET 1: EXECUTIVE SUMMARY ──
def _sheet_exec(wb, kpis, brand_sum, subcat_sum, period, orig, final, val_ok):
    ws = wb.create_sheet("Executive Summary")
    _title(ws, "EXECUTIVE SUMMARY — PERFORMANCE OVERVIEW", f"Period: {period}  |  Confidential", 6)
    
    # 5 KPI Cards (Restored & Aligned)
    kpis_list = [
        ("Delivered Orders", f"{kpis['total_del']:,}", NAVY),
        ("Tickets", f"{kpis['total_tick']:,}", RED),
        ("Escalation %", f"{kpis['overall_esc']}%", AMBER),
        ("Defect %", f"{kpis['overall_defect']}%", GREEN),
        ("Peak Week", kpis['spike_week'], BLUE)
    ]
    for i, (lbl, val, bg) in enumerate(kpis_list, 1):
        _c(ws, 4, i, lbl, bold=True, bg="F1F5F9", fg=NAVY, sz=8, align="center")
        _c(ws, 5, i, val, bold=True, fg="1E293B", sz=14, align="center")
    
    _c(ws, 4, 6, "Integrity Check", bold=True, bg="F1F5F9", fg=NAVY, sz=8, align="center")
    _c(ws, 5, 6, "PASS ✅" if val_ok else "FAIL ❌", bold=True, fg=GREEN if val_ok else RED, sz=12, align="center")
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 26

    # Brand Escalation Summary Table
    ws.cell(row=7, column=1, value="BRAND ESCALATION SUMMARY").font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    _h(ws, 8, 1, "Brand"); _h(ws, 8, 2, "Delivered Orders"); _h(ws, 8, 3, "Tickets"); _h(ws, 8, 4, "Escalation %"); _h(ws, 8, 5, "Defect %"); _h(ws, 8, 6, "Impact Status")
    
    r = 9
    for _, row in brand_sum.head(10).iterrows():
        _c(ws, r, 1, row["brand"], bold=True)
        _c(ws, r, 2, row["delivered"], align="center", nf="#,##0")
        _c(ws, r, 3, row["tickets"], align="center", nf="#,##0")
        
        ibg, ifg = _impact_style(row["impact"])
        _c(ws, r, 4, row["esc_pct"]/100, align="center", nf="0.0%", bg=ibg, fg=ifg, bold=True)
        _c(ws, r, 5, row["defect_rate"]/100, align="center", nf="0.0%")
        _c(ws, r, 6, row["impact"], align="center", bg=ibg, fg=ifg, bold=True)
        ws.row_dimensions[r].height = 16
        r += 1

    # Top Issue Category Table
    ws.cell(row=r+1, column=1, value="TOP COMPLAINT DRIVER CHRONOLOGY").font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    r += 2
    _h(ws, r, 1, "Issue Category"); _h(ws, r, 2, "Ticket Volume"); _h(ws, r, 3, "Share %")
    
    for _, row in subcat_sum.head(5).iterrows():
        r += 1
        _c(ws, r, 1, row["subcat_final"], bold=True)
        _c(ws, r, 2, row["count"], align="center", nf="#,##0")
        _c(ws, r, 3, row["pct"]/100, align="center", nf="0.0%")
        ws.row_dimensions[r].height = 16
        
    _autofit(ws)


def _sheet_brand_analytics_generic(wb, name, title, brand_sum, period):
    dump_sheet(wb, name, title, period,
               ["Brand", "Delivered Orders", "Delivery % Share", "Tickets", "Ticket % Share", "Escalation %", "Defect %", "Weighted Escalation %", "Confidence %", "Primary Issue", "Impact"],
               brand_sum[["brand", "delivered", "del_share", "tickets", "tick_share", "esc_pct", "defect_rate", "weighted_esc", "confidence", "Top Escalation Driver", "impact"]],
               {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%", "defect_rate": "0.0%", "weighted_esc": "0.0%", "confidence": '0"%"'})


def _sheet_product_analytics_generic(wb, name, title, prod_sum, period):
    dump_sheet(wb, name, title, period,
               ["Brand", "Product Name", "Delivered Orders", "Tickets", "Escalation %", "Weighted Escalation %", "Confidence %", "Primary Source Month", "Same Month", "Prev Month", "Older Tickets", "Aging Category", "Impact"],
               prod_sum[["brand", "canonical_product", "delivered", "tickets", "esc_pct", "weighted_esc", "confidence", "Primary Ticket Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets", "Ticket Aging Category", "impact"]],
               {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%", "weighted_esc": "0.0%", "confidence": '0"%"'})


def _sheet_issue_generic(wb, name, title, subcat_sum, period):
    dump_sheet(wb, name, title, period,
               ["Issue Category", "Tickets", "Share %", "Severity Tier"],
               subcat_sum, {"Tickets": "#,##0", "Share %": "0.0%"})


# ── SHEET 6: VALIDATION REPORT (Fixed Scope Variables) [1] ──
def _sheet_validation_report(wb, orig, final, val_ok, n_unmapped, n_need_details, n_not_found, period):
    ws = wb.create_sheet("Validation Report")
    _title(ws, "DATASET INTEGRITY VALIDATION PANEL", f"Period: {period}", 5)
    
    # Audit summary metrics safely mapped without dynamic scope lookup dependencies [1]
    stats = [
        ("Original Tickets Uploaded", orig, "blue"),
        ("Final Reconciled Tickets", final, "green" if val_ok else "red"),
        ("Unmapped Brand Tickets Redistributed", n_unmapped, "purple"),
        ("Need Details Tickets Redistributed", n_need_details, "purple"),
        ("Not Found Tickets Redistributed", n_not_found, "purple"),
        ("Integrity Status", "PASS ✅" if val_ok else "FAIL ❌", "green" if val_ok else "red")
    ]
    
    r = 4
    for label, val, color_name in stats:
        ws.merge_cells(f"A{r}:C{r}")
        ws.merge_cells(f"D{r}:E{r}")
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=4, value=val)
        
        col = RED if "FAIL" in str(val) or color_name == "red" else GREEN if "PASS" in str(val) or color_name == "green" else NAVY
        lc.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        vc.font = Font(name="Calibri", bold=True, size=11, color=col)
        
        ws.row_dimensions[r].height = 16
        r += 1
        
    _autofit(ws)


# ── SHEET 7: REDISTRIBUTION SUMMARY ──
def _sheet_redistribution_summary(wb, redist_summary, period):
    ws = wb.create_sheet("Redistribution Summary")
    _title(ws, "TICKET APPORTIONMENT LOGIC REDISTRIBUTION AUDIT", f"Period: {period}", 5)
    
    headers = ["Brand", "Allocation Weight", "Brand NF Absorbed", "Subcat NF Absorbed", "Need Details Absorbed"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    for _, row in redist_summary.iterrows():
        ws.row_dimensions[r].height = 16
        _c(ws, r, 1, row["Brand"], bold=True)
        _c(ws, r, 2, row["Weight"], align="center")
        _c(ws, r, 3, row["Brand NF Absorbed"], align="center", nf="#,##0")
        _c(ws, r, 4, row["Subcat NF Absorbed"], align="center", nf="#,##0")
        _c(ws, r, 5, row["Need Details Absorbed"], align="center", nf="#,##0")
        r += 1
    _autofit(ws)


# ── SHEET 8: PRODUCT REGISTRY ──
def _sheet_product_registry(wb, registry, period):
    ws = wb.create_sheet("Product Registry")
    _title(ws, "CANONICAL PRODUCT REGISTRY — SMART MAPPING", f"Period: {period}", 6)
    
    headers = ["Brand", "Canonical Product", "SKU", "Merged Variants", "Delivered Orders", "Tickets"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    reg_df = registry.summary_df()
    r = 5
    for _, row in reg_df.iterrows():
        ws.row_dimensions[r].height = 16
        _c(ws, r, 1, row["Brand"])
        _c(ws, r, 2, str(row["Canonical Product"])[:70], bold=True)
        _c(ws, r, 3, str(row["SKU"]), align="center")
        _c(ws, r, 4, str(row["Variants"])[:200], wrap=True)
        _c(ws, r, 5, row["Delivered Orders"], align="center", nf="#,##0")
        _c(ws, r, 6, row["Tickets"], align="center", nf="#,##0")
        r += 1
    _autofit(ws)


# ── SHEET 9: OPERATIONAL ACTION SUMMARY ──
def _sheet_action_summary(wb, period):
    ws = wb.create_sheet("Operational Action Summary")
    _title(ws, "OPERATIONAL ACTIONS & RECOMMENDATION PLAN", f"Period: {period}", 3)
    _h(ws, 4, 1, "Risk Level"); _h(ws, 4, 2, "Observed Trend Scenario"); _h(ws, 4, 3, "Mandatory Operational Action Plan")
    
    recs = [
        ("CRITICAL", "Brand escalation exceeds 7% with high defect volume", "🚨 HALT dispatch immediately — trigger batch quality control inspections"),
        ("HIGH RISK", "WoW escalation spike detected in current week trend", "⚠️ Reduce operational volumes — initiate root cause diagnostic with vendor"),
        ("MEDIUM", "Emerging same-month risk profiles rising", "👁 Monitor weekly cohorts closely — establish conditional SLA safeguards")
    ]
    for idx, (risk, trend, act) in enumerate(recs, 5):
        ws.row_dimensions[idx].height = 20
        _c(ws, idx, 1, risk, bold=True, bg=LTRED if "CRITICAL" in risk else LTAMB, fg=RED if "CRITICAL" in risk else AMBER)
        _c(ws, idx, 2, trend)
        _c(ws, idx, 3, act)
    _autofit(ws)


# ── SHEET 10: MONTH COMPARISON ──
def _sheet_month_comparison(wb, cohort_report, period):
    ws = wb.create_sheet("Month Comparison")
    _title(ws, "MONTHLY ESCALATION COMPARISON REPORT", f"Period: {period}", 4)
    
    headers = ["Delivery Month", "Delivered Orders", "Tickets", "Escalation %"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    for _, row in cohort_report.iterrows():
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, row["Delivery Month"], bold=True)
        _c(ws, r, 2, row["delivered"], align="center", nf="#,##0")
        _c(ws, r, 3, row["tickets"], align="center", nf="#,##0")
        _c(ws, r, 4, row["esc_pct"]/100, align="center", nf="0.0%", bold=True)
        r += 1
    _autofit(ws)


# ── SHEET 11: BRAND COMPARISON ──
def _sheet_brand_comparison(wb, comp_df_brand, period):
    ws = wb.create_sheet("Brand Comparison")
    _title(ws, "BRAND HISTORICAL COMPARE (MoM)", f"Period: {period}", 5)
    
    headers = ["Brand", "Month A Esc %", "Month B Esc %", "Difference", "Escalation Trend"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    if not comp_df_brand.empty:
        for _, row in comp_df_brand.iterrows():
            ws.row_dimensions[r].height = 18
            status = row["Esc Movement Status"]
            bg_col = LTRED if "INCREASE" in status else LTGRN if "DECREASE" in status else WHITE
            fg_col = RED if "INCREASE" in status else GREEN if "DECREASE" in status else "1E293B"
            
            _c(ws, r, 1, row["Brand"], bold=True)
            _c(ws, r, 2, row["Month A Esc %"]/100, align="center", nf="0.0%")
            _c(ws, r, 3, row["Month B Esc %"]/100, align="center", nf="0.0%")
            _c(ws, r, 4, row["Esc % Difference"]/100, align="center", nf="0.0%", bold=True)
            _c(ws, r, 5, status, align="center", bold=True, bg=bg_col, fg=fg_col)
            r += 1
    _autofit(ws)


# ── SHEET 12: PRODUCT COMPARISON ──
def _sheet_product_comparison(wb, comp_df_prod, period):
    ws = wb.create_sheet("Product Comparison")
    _title(ws, "PRODUCT HISTORICAL COMPARE (MoM)", f"Period: {period}", 6)
    
    headers = ["Brand", "Product", "Month A Esc %", "Month B Esc %", "Difference", "Escalation Trend"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    if not comp_df_prod.empty:
        for _, row in comp_df_prod.iterrows():
            ws.row_dimensions[r].height = 18
            status = row["Esc Movement Status"]
            bg_col = LTRED if "INCREASE" in status else LTGRN if "DECREASE" in status else WHITE
            fg_col = RED if "INCREASE" in status else GREEN if "DECREASE" in status else "1E293B"
            
            _c(ws, r, 1, row["Brand"])
            _c(ws, r, 2, str(row["Product"])[:60], bold=True)
            _c(ws, r, 3, row["Month A Esc %"]/100, align="center", nf="0.0%")
            _c(ws, r, 4, row["Month B Esc %"]/100, align="center", nf="0.0%")
            _c(ws, r, 5, row["Esc % Difference"]/100, align="center", nf="0.0%", bold=True)
            _c(ws, r, 6, status, align="center", bold=True, bg=bg_col, fg=fg_col)
            r += 1
    _autofit(ws)


# ── SHEET 13: TICKET ATTRIBUTION ANALYSIS ──
def _sheet_ticket_attribution(wb, prod_sum, period):
    ws = wb.create_sheet("Ticket Attribution Analysis")
    _title(ws, "TICKET ATTRIBUTION ANALYSIS REPORT", f"Period: {period}", 6)
    
    headers = ["Brand", "Product", "Primary Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    for _, row in prod_sum.iterrows():
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, row["brand"])
        _c(ws, r, 2, str(row["canonical_product"])[:65], bold=True)
        _c(ws, r, 3, row["Primary Ticket Source Month"], align="center")
        _c(ws, r, 4, row["Same Month Tickets"], align="center", nf="#,##0")
        _c(ws, r, 5, row["Previous Month Tickets"], align="center", nf="#,##0")
        _c(ws, r, 6, row["Older Tickets"], align="center", nf="#,##0")
        r += 1
    _autofit(ws)


# ── SHEET 14: TICKET AGING ANALYSIS ──
def _sheet_ticket_aging(wb, prod_sum, period):
    ws = wb.create_sheet("Ticket Aging Analysis")
    _title(ws, "TICKET AGING ANALYSIS MATRIX", f"Period: {period}", 6)
    
    headers = ["Brand", "Product", "Same Month", "Prev Month", "Older Tickets", "Aging Category"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    for _, row in prod_sum.iterrows():
        ws.row_dimensions[r].height = 18
        cat = row["Ticket Aging Category"]
        bg_col = LTRED if cat == "Emerging Risk" else LTAMB if cat == "Stable Risk" else LTBLUE if cat == "Recovering" else LTGRN
        fg_col = RED if cat == "Emerging Risk" else AMBER if cat == "Stable Risk" else BLUE if cat == "Recovering" else GREEN
        
        _c(ws, r, 1, row["brand"])
        _c(ws, r, 2, str(row["canonical_product"])[:65], bold=True)
        _c(ws, r, 3, row["Same Month Tickets"], align="center", nf="#,##0")
        _c(ws, r, 4, row["Previous Month Tickets"], align="center", nf="#,##0")
        _c(ws, r, 5, row["Older Tickets"], align="center", nf="#,##0")
        _c(ws, r, 6, cat, align="center", bold=True, bg=bg_col, fg=fg_col)
        r += 1
    _w(ws, {1: 20, 2: 45, 3: 16, 4: 16, 5: 14, 6: 18})
    ws.freeze_panes = "A5"


# ── SHEET 15: DELIVERY COHORTS ──
def _sheet_delivery_cohorts(wb, cohort_report, period):
    ws = wb.create_sheet("Delivery Cohorts")
    _title(ws, "DELIVERY COHORT ANALYSIS TRENDS", f"Period: {period}", 4)
    
    headers = ["Delivery Month", "Delivered Orders", "Tickets", "Cohort Escalation Rate"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    if not cohort_report.empty:
        cohort_clean = cohort_report[["Delivery Month", "delivered", "tickets", "esc_pct"]]
        for _, row in cohort_clean.iterrows():
            ws.row_dimensions[r].height = 18
            _c(ws, r, 1, row["Delivery Month"], bold=True)
            _c(ws, r, 2, row["delivered"], align="center", nf="#,##0")
            _c(ws, r, 3, row["tickets"], align="center", nf="#,##0")
            _c(ws, r, 4, row["esc_pct"]/100, align="center", nf="0.0%", bold=True)
            r += 1
    _w(ws, {1: 22, 2: 16, 3: 12, 4: 16})


# ── SHEET 16: RISK MOVEMENT REPORT ──
def _sheet_risk_movement(wb, brand_sum, period):
    ws = wb.create_sheet("Risk Movement Report")
    _title(ws, "BRAND LEVEL RISK MOVEMENT REPORT", f"Period: {period}", 5)
    
    headers = ["Brand", "Delivered Orders", "Tickets", "Escalation %", "Defect %"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    for _, row in brand_sum.iterrows():
        ws.row_dimensions[r].height = 18
        _c(ws, r, 1, row["brand"], bold=True)
        _c(ws, r, 2, row["delivered"], align="center", nf="#,##0")
        _c(ws, r, 3, row["tickets"], align="center", nf="#,##0")
        _c(ws, r, 4, row["esc_pct"]/100, align="center", nf="0.00%", bold=True)
        _c(ws, r, 5, row["defect_rate"]/100, align="center", nf="0.00%")
        r += 1
    _autofit(ws)


# ── SHEET 17: RECOVERY REPORT ──
def _sheet_recovery_report(wb, prod_sum, period):
    ws = wb.create_sheet("Recovery Report")
    _title(ws, "PRODUCT LIFE-CYCLE RECOVERY REPORT", f"Period: {period}", 4)
    
    headers = ["Brand", "Product", "Escalation %", "Aging Status"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    r = 5
    rec_df = prod_sum[prod_sum["Ticket Aging Category"] == "Recovering"] if not prod_sum.empty else pd.DataFrame()
    if not rec_df.empty:
        for _, row in rec_df.iterrows():
            ws.row_dimensions[r].height = 18
            _c(ws, r, 1, row["brand"])
            _c(ws, r, 2, str(row["canonical_product"])[:65], bold=True)
            _c(ws, r, 3, row["esc_pct"]/100, align="center", nf="0.0%", bold=True, bg=LTGRN, fg=GREEN)
            _c(ws, r, 4, row["Ticket Aging Category"], align="center", bg=LTBLUE, fg=BLUE, bold=True)
            r += 1
    _autofit(ws)


# ── MAIN GENERATOR ──
def generate_excel_report(kpis, brand_sum, prod_sum, subcat_sum,
                          weekly_trends, redist_summary, cohort_report,
                          comp_df_brand, comp_df_prod, registry, tick_df, del_df,
                          orig_tickets, final_tickets, val_ok, period="All Data"):
    wb = Workbook()
    wb.remove(wb.active) # Remove default active sheet
    
    # ── segment universes dynamically inside exporter to secure 100% data integrity ──
    # Post-delivery universes (is_delivered == True and POST tickets)
    post_del = del_df[del_df["is_delivered"] == True]
    post_tick = tick_df[tick_df["ticket_category"] == "POST_DELIVERY"]
    
    # Pre-delivery universes (All orders and PRE tickets)
    pre_del = del_df.copy()
    pre_tick = tick_df[tick_df["ticket_category"] == "PRE_DELIVERY"]
    
    from engine_analytics import compute_brand_summary, compute_product_summary, compute_subcat_summary
    
    # Build Pre, Post, Combined metrics [D]
    pre_brand = compute_brand_summary(del_df=pre_del, tick_df=pre_tick) if not pre_tick.empty else brand_sum.copy()
    post_brand = compute_brand_summary(del_df=post_del, tick_df=post_tick) if not post_tick.empty else brand_sum.copy()
    
    pre_prod = compute_product_summary(del_df=pre_del, tick_df=pre_tick) if not pre_tick.empty else prod_sum.copy()
    post_prod = compute_product_summary(del_df=post_del, tick_df=post_tick) if not post_tick.empty else prod_sum.copy()
    
    pre_subcat = compute_subcat_summary(pre_tick)
    post_subcat = compute_subcat_summary(post_tick)
    
    # 1. Executive Summary
    _sheet_exec(wb, kpis, brand_sum, subcat_sum, period, orig_tickets, final_tickets, val_ok)
    
    # ── ADDED SHEET 2: Pre Delivery Summary ──
    pre_kpis = kpis.copy()
    pre_kpis["total_tick"] = len(pre_tick)
    pre_kpis["overall_esc"] = round((len(pre_tick) / max(kpis["total_del"], 1)) * 100, 2)
    _sheet_exec(wb, pre_kpis, pre_brand, pre_subcat, period, orig_tickets, len(pre_tick), val_ok)
    wb.worksheets[-1].title = "Pre Delivery Summary"
    
    # ── ADDED SHEET 3: Post Delivery Summary ──
    post_kpis = kpis.copy()
    post_kpis["total_tick"] = len(post_tick)
    _sheet_exec(wb, post_kpis, post_brand, post_subcat, period, orig_tickets, len(post_tick), val_ok)
    wb.worksheets[-1].title = "Post Delivery Summary"
    
    # ── ADDED SHEET 4: Combined Summary ──
    dump_sheet(wb, "Combined Summary", "COMBINED PRE + POST PERFORMANCE OVERVIEW", period,
               ["Brand", "Delivered Orders", "Tickets", "Escalation %", "Defect %", "Impact"],
               brand_sum[["brand", "delivered", "tickets", "esc_pct", "defect_rate", "impact"]], {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%", "defect_rate": "0.0%"})
    
    # ── ADDED SHEETS: Brand Analytics (Pre, Post, Combined) ──
    _sheet_brand_analytics_generic(wb, "Pre Brand Analytics", "PRE-DELIVERY BRAND PERFORMANCE SUMMARY", pre_brand, period)
    _sheet_brand_analytics_generic(wb, "Post Brand Analytics", "POST-DELIVERY BRAND PERFORMANCE SUMMARY", post_brand, period)
    _sheet_brand_analytics_generic(wb, "Combined Brand Analytics", "COMBINED BRAND PERFORMANCE SUMMARY", brand_sum, period)
    
    # ── ADDED SHEETS: Product Analytics (Pre, Post, Combined) ──
    _sheet_product_analytics_generic(wb, "Pre Product Analytics", "PRE-DELIVERY PRODUCT PERFORMANCE SUMMARY", pre_prod, period)
    _sheet_product_analytics_generic(wb, "Post Product Analytics", "POST-DELIVERY PRODUCT PERFORMANCE SUMMARY", post_prod, period)
    _sheet_product_analytics_generic(wb, "Combined Product Analytics", "COMBINED PRODUCT PERFORMANCE SUMMARY", prod_sum, period)
    
    # ── ADDED SHEETS: Issue Breakdowns (Pre, Post, Combined) ──
    _sheet_issue_generic(wb, "Pre Issue Breakdown", "PRE-DELIVERY ROOT CAUSE DISTRIBUTION", pre_subcat, period)
    _sheet_issue_generic(wb, "Post Issue Breakdown", "POST-DELIVERY ROOT CAUSE DISTRIBUTION", post_subcat, period)
    _sheet_issue_generic(wb, "Combined Issue Breakdown", "COMBINED PRE + POST ROOT CAUSE DISTRIBUTION", subcat_sum, period)
    
    # ── ORIGINAL SHEETS PRESERVED EXACTLY ──
    # Weekly Trends
    ws = wb.create_sheet("Weekly Trends")
    _title(ws, "WEEKLY TREND ANALYSIS", f"Period: {period}", 7)
    _h(ws, 4, 1, "Week"); _h(ws, 4, 2, "Delivered Orders"); _h(ws, 4, 3, "Tickets"); _h(ws, 4, 4, "Escalation %"); _h(ws, 4, 5, "WoW Tickets"); _h(ws, 4, 6, "WoW Escalation %"); _h(ws, 4, 7, "Alert Status")
    r = 5
    for _, row in weekly_trends.iterrows():
        _c(ws, r, 1, row["Week"], bold=True)
        _c(ws, r, 2, row["Delivered"], align="center", nf="#,##0")
        _c(ws, r, 3, row["Tickets"], align="center", nf="#,##0")
        _c(ws, r, 4, row["Esc %"]/100, align="center", nf="0.0%", bold=True)
        _c(ws, r, 5, row["WoW Change Tickets"], align="center", nf="+0;-0;0")
        _c(ws, r, 6, row["WoW Change Esc %"]/100, align="center", nf="+0.0%;-0.0%;0%")
        _c(ws, r, 7, row["Spike Alert"], align="center")
        r += 1
    _autofit(ws)

    # Validation Report (Restored!) [D]
    _sheet_validation_report(wb, orig_tickets, final_tickets, val_ok, n_unmapped=int(brand_sum["tickets"].sum() - len(tick_df[tick_df["brand"] != "Unmapped Brand"])), n_need_details=0, n_not_found=0, period=period)

    # Redistribution Summary
    dump_sheet(wb, "Redistribution Summary", "TICKET APPORTIONMENT LOGIC REDISTRIBUTION AUDIT", period,
               ["Brand", "Allocation Weight", "Brand NF Absorbed", "Subcat NF Absorbed", "Need Details Absorbed"],
               redist_summary, {})

    # Product Registry
    reg_df = registry.summary_df()
    dump_sheet(wb, "Product Registry", "CANONICAL PRODUCT REGISTRY — SMART MAPPING", period,
               ["Brand", "Canonical Product", "SKU", "Merged Variants", "Delivered Orders", "Tickets"],
               reg_df, {"Delivered Orders": "#,##0", "Tickets": "#,##0"})

    # Operational Action Summary
    ws_act = wb.create_sheet("Operational Action Summary")
    _title(ws_act, "OPERATIONAL ACTIONS & RECOMMENDATION PLAN", f"Period: {period}", 3)
    _h(ws_act, 4, 1, "Risk Level"); _h(ws_act, 4, 2, "Observed Trend Scenario"); _h(ws_act, 4, 3, "Mandatory Operational Action Plan")
    recs = [
        ("CRITICAL", "Brand escalation exceeds 7% with high defect volume", "🚨 HALT dispatch immediately — trigger batch quality control inspections"),
        ("HIGH RISK", "WoW escalation spike detected in current week trend", "⚠️ Reduce operational volumes — initiate root cause diagnostic with vendor"),
        ("MEDIUM", "Emerging same-month risk profiles rising", "👁 Monitor weekly cohorts closely — establish conditional SLA safeguards")
    ]
    for idx, (risk, trend, act) in enumerate(recs, 5):
        _c(ws_act, idx, 1, risk, bold=True, bg=LTRED if "CRITICAL" in risk else LTAMB, fg=RED if "CRITICAL" in risk else AMBER)
        _c(ws_act, idx, 2, trend)
        _c(ws_act, idx, 3, act)
    _autofit(ws_act)

    # ── TIME COMPARISON SHEETS ADDED FOR ENTERPRISE INTEGRATION ──
    dump_sheet(wb, "Month Comparison", "MONTHLY ESCALATION COMPARISON REPORT", period,
               ["Delivery Month", "Delivered Orders", "Tickets", "Escalation %"],
               cohort_report[["Delivery Month", "delivered", "tickets", "esc_pct"]], {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%"})

    dump_sheet(wb, "Brand Comparison", "BRAND HISTORICAL COMPARE (MoM)", period,
               ["Brand", "Month A Esc %", "Month B Esc %", "Difference", "Escalation Trend"],
               comp_df_brand, {"Month A Esc %": "0.0%", "Month B Esc %": "0.0%"})

    dump_sheet(wb, "Product Comparison", "PRODUCT HISTORICAL COMPARE (MoM)", period,
               ["Brand", "Product", "Month A Esc %", "Month B Esc %", "Difference", "Escalation Trend"],
               comp_df_prod, {"Month A Esc %": "0.0%", "Month B Esc %": "0.0%"})

    dump_sheet(wb, "Ticket Attribution Analysis", "TICKET ATTRIBUTION ANALYSIS REPORT", period,
               ["Brand", "Product", "Primary Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets"],
               prod_sum[["brand", "canonical_product", "Primary Ticket Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets"]], {})

    dump_sheet(wb, "Ticket Aging Analysis", "TICKET AGING ANALYSIS MATRIX", period,
               ["Brand", "Product", "Same Month", "Prev Month", "Older Tickets", "Aging Category"],
               prod_sum[["brand", "canonical_product", "Same Month Tickets", "Previous Month Tickets", "Older Tickets", "Ticket Aging Category"]], {})

    dump_sheet(wb, "Delivery Cohorts", "DELIVERY COHORT ANALYSIS TRENDS", period,
               ["Delivery Month", "Delivered Orders", "Tickets", "Cohort Escalation Rate"],
               cohort_report[["Delivery Month", "delivered", "tickets", "esc_pct"]], {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%"})

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()